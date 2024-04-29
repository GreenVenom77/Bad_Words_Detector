import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO


class ExcelPlotter:
    def __init__(self, excel_file_path, sheet1_name, sheet2_name):
        self.excel_file_path = excel_file_path
        self.sheet1_name = sheet1_name
        self.sheet2_name = sheet2_name
        self.col_names = [
            "avg_Reading Time",
            "Total Reading Time",
            "Max Reading Time",
            "Min Reading Time",
            "avg_filtering Time",
            "Total_filtering Time",
            "Max_filtering Time",
            "Min_filtering Time",
            "Avg Frame_processing_Time",
            "Total Frame_processing_Time",
            "Max Frame_processing_Time",
            "Min Frame_processing_Time",
        ]
        self.sns_style = "whitegrid"

    def read_excel_data(self):
        self.sheet1_df = pd.read_excel(
            self.excel_file_path, sheet_name=self.sheet1_name
        )
        self.sheet2_df = pd.read_excel(
            self.excel_file_path, sheet_name=self.sheet2_name
        )
        self.concatenated_df = pd.concat(
            [self.sheet1_df, self.sheet2_df], keys=[self.sheet1_name, self.sheet2_name]
        )

    def plot_and_embed(self):
        wb = load_workbook(self.excel_file_path)
        if "BarChart_1" not in wb.sheetnames:
            ws = wb.create_sheet(title="BarChart_1")
        else:
            ws = wb["BarChart_1"]

        for i, col in enumerate(self.col_names, start=1):
            plt.figure(figsize=(11, 4))
            sns.set_style(self.sns_style)
            sns.barplot(
                x="D.frame size",
                y=col,
                hue=self.concatenated_df.index.get_level_values(0),
                data=self.concatenated_df.reset_index(),
            )
            plt.title(f"{col} Comparison of Algorithms by Chunk Size")
            plt.xlabel("Chunk Size")
            plt.ylabel(f"{col}")
            plt.legend(title="Algorithm")

            buffer = BytesIO()
            plt.savefig(buffer, format="png")
            buffer.seek(0)

            img = Image(buffer)
            ws.add_image(img, f"A{(i - 1) * 22 + 1}")
            plt.close()

        wb.save(self.excel_file_path)

    def plot_ahocorasick_only(self):
        ahocorasick_df = self.sheet1_df
        self._plot_single_algorithm(ahocorasick_df, "AhoCorasick", "A", "#1F77B4")

    def plot_regex_only(self):
        regex_df = self.sheet2_df
        self._plot_single_algorithm(regex_df, "Regex", "Q", "#FF7F0E")

    def _plot_single_algorithm(self, df, algorithm_name, target_class, color_bar):
        wb = load_workbook(self.excel_file_path)
        if "BarChart_2" not in wb.sheetnames:
            ws = wb.create_sheet(title="BarChart_2")
        else:
            ws = wb["BarChart_2"]

        for i, col in enumerate(self.col_names, start=1):
            plt.figure(figsize=(10, 4))
            sns.set_style(self.sns_style)
            sns.barplot(x="D.frame size", y=col, data=df, color=color_bar)
            plt.title(f"{col} Comparison using {algorithm_name} by Chunk Size")
            plt.xlabel("Chunk Size")
            plt.ylabel(f"{col}")
            plt.legend([algorithm_name])

            buffer = BytesIO()
            plt.savefig(buffer, format="png")
            buffer.seek(0)

            img = Image(buffer)
            ws.add_image(img, f"{target_class}{(i - 1) * 22 + 1}")
            plt.close()

        wb.save(self.excel_file_path)


# Usage
excel_plotter = ExcelPlotter(
    "./output/StatisticalSummary_CS342Spring2024.xlsx", "AhoCorasick", "Regex"
)
excel_plotter.read_excel_data()
excel_plotter.plot_and_embed()
excel_plotter.plot_ahocorasick_only()
excel_plotter.plot_regex_only()
print("The program finished successfully")
