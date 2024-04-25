import logging
import os
import threading

import openpyxl
from pandas import DataFrame, concat

from arguments import Args
from chunks_processing_info import ChunkInfo, elapsed


class StatisticsWriter:
    def __init__(self, args: Args) -> None:
        self.args = args

    def convert_to_data_frame(self, chunks_info: list[ChunkInfo]) -> DataFrame:
        return DataFrame(
            [
                {
                    "chunk_size": self.args.chunk_size,
                    "chunk_number": i,
                    "healthy_records": chunk.number_of_healthy,
                    "unhealthy_records": chunk.number_of_unhealthy,
                    "reading_time": chunk.reading_time,
                    "filtering_time": chunk.filtering_time,
                    "frame_total_time": round(
                        chunk.reading_time + chunk.filtering_time,
                        self.args.rounding_place,
                    ),
                }
                for i, chunk in enumerate(chunks_info, start=1)
            ]
        )

    def calculate_aggregation_values(
        self, df: DataFrame
    ) -> dict[str, dict[str, float | int]]:
        agg_values = dict[str, dict[str, float | int]]()
        for column in df.columns[2:4]:
            agg_values[column] = {
                "sum": round(df[column].sum(), self.args.rounding_place)
            }
        for column in df.columns[4:]:
            agg_values[column] = {
                "sum": round(df[column].sum(), self.args.rounding_place),
                "average": round(df[column].mean(), self.args.rounding_place),
                "max": df[column].max(),
                "min": df[column].min(),
            }
        return agg_values

    def write_csv(
        self, df: DataFrame, aggregation_values: dict[str, dict[str, float | int]]
    ):
        aggregation_rows = list[list[str | None]]()
        for key in ["sum", "average", "max", "min"]:
            label = key if key != "sum" else "total"
            fields = len(df.columns) - sum(
                key in dict for dict in aggregation_values.values()
            )
            aggregation_rows.append(
                [
                    None,
                ]
                * fields
                + [
                    f"{label}:{aggregation_values[column][key]}"
                    for column in df.columns[fields:]
                ]
            )
        csv_df = concat(
            [df, DataFrame(aggregation_rows, columns=df.columns, index=None)], axis=0
        )
        # replace csv columns
        csv_df.columns = [
            "chunk_size",
            "chunk_number",
            "No of healthy records",
            "No of unhealthy records",
            "reading_time",
            "filtering_time",
            "frame_Total_Time",
        ]
        csv_df.to_csv(
            path_or_buf=f"output/Detailed_Log_{self.args.filter_mode.name}.csv",
            mode="a",
            index=False,
        )

    def write_excel(self, aggregations: dict[str, dict[str, float | int]]):
        path = "output/StatisticalSummary_CS342Spring2024.xlsx"
        # setup row
        data: dict[str, int | float] = {
            "D.frame size": self.args.chunk_size,
            "Sum No of healthy records": aggregations["healthy_records"]["sum"],
            "Sum No of unhealthy records": aggregations["unhealthy_records"]["sum"],
            "avg_Reading Time": aggregations["reading_time"]["average"],
            "Total Reading Time": aggregations["reading_time"]["sum"],
            "Max Reading Time": aggregations["reading_time"]["max"],
            "Min Reading Time": aggregations["reading_time"]["min"],
            "avg_filtering Time": aggregations["filtering_time"]["average"],
            "Total_filtering Time": aggregations["filtering_time"]["sum"],
            "Max_filtering Time": aggregations["filtering_time"]["max"],
            "Min_filtering Time": aggregations["filtering_time"]["min"],
            "Avg Frame_processing_Time": aggregations["frame_total_time"]["average"],
            "Total Frame_processing_Time": aggregations["frame_total_time"]["sum"],
            "Max Frame_processing_Time": aggregations["frame_total_time"]["max"],
            "Min Frame_processing_Time": aggregations["frame_total_time"]["min"],
        }
        columns = list(data.keys())
        values = list(data.values())
        # check if the file exists
        if not os.path.exists(path):
            # create a new workbook and worksheet if the file doesn't exist
            workbook = openpyxl.Workbook()
            worksheet = workbook.create_sheet(self.args.filter_mode.name)
            worksheet.append(columns)
            # if excel created a default sheet
            # if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
        else:
            workbook = openpyxl.load_workbook(path)
            if self.args.filter_mode.name in workbook.sheetnames:
                worksheet = workbook[self.args.filter_mode.name]
            else:
                worksheet = workbook.create_sheet(self.args.filter_mode.name)
                worksheet.append(columns)

        # append data to the worksheet
        worksheet.append(values)
        # save the workbook
        workbook.save(path)

    def start(self, chunks_info: list[ChunkInfo]) -> None:
        logging.info(
            "%s  Writer:starting write excel and csv files statistics ....",
            elapsed(self.args.starting_time),
        )
        df = self.convert_to_data_frame(chunks_info)
        aggregation_values = self.calculate_aggregation_values(df)
        if not os.path.exists("output"):
            os.makedirs("output")
        csv_thread = threading.Thread(
            target=lambda: self.write_csv(df, aggregation_values)
        )
        excel_thread = threading.Thread(
            target=lambda: self.write_excel(aggregation_values)
        )
        excel_thread.start()
        csv_thread.start()
        csv_thread.join()
        excel_thread.join()

        logging.info(
            "%s  Writer:finish of write excel and csv files statistics .",
            elapsed(self.args.starting_time),
        )
