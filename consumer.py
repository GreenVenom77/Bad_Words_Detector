import multiprocessing.process
import multiprocessing.queues
from queue import Queue
import threading
import time
import openpyxl
from functools import reduce
import os
import logging
import pandas as pd
import csv
from filter import TextFilter
from time_helper import elapsed

logging.basicConfig(
    filename="logfile.log",
    format=" %(asctime)s %(levelname)-8s %(message)s",
    level=logging.INFO,
)


class DummyLock:
    def acquire(self):
        pass

    def release(self):
        pass


# cpu bounds
class Consumer:
    def __init__(
        self,
        text_filter: TextFilter,
        input_queue: Queue[tuple[int, pd.DataFrame]],
        time_dict,
        use_time_dict_lock: bool,
    ):
        self.__text_filter = text_filter
        self.__input_queue = input_queue
        self.__time_dict = time_dict
        if use_time_dict_lock:
            self.__time_dict_lock = multiprocessing.Lock()
        else:
            self.__time_dict_lock = DummyLock()

    def write_csv_statistics(self, filename):
        with open("output/" + filename, "a", newline="") as csvfile:
            fieldnames = [
                "chunk_size",
                "chunk_number",
                "reading_time",
                "filtering_time",
                "writing_time",
            ]
            writer = csv.DictWriter(csvfile, fieldnames)
            writer.writeheader()

            for i in range(len(self.__time_dict["reading"])):
                row = {
                    "chunk_size": self.__time_dict["chunksize"],
                    "chunk_number": i + 1,
                    "reading_time": self.__time_dict["reading"][i],
                    "filtering_time": self.__time_dict["filtering"][i],
                    "writing_time": self.__time_dict["writing"][i],
                }
                writer.writerow(row)

    def excel_writer(self, sheet_name):

        # measure reading total time and avg time
        total_time_of_read = sum(self.__time_dict["reading"])
        avg_time_of_read = total_time_of_read / self.__time_dict["number of chunks"]

        # measure filtering total time and avg time
        total_time_of_filter = sum(self.__time_dict["filtering"])
        avg_time_of_filter = total_time_of_filter / self.__time_dict["number of chunks"]
        # measure writing total time and avg time
        total_time_of_writing = sum(self.__time_dict["writing"])
        avg_time_of_writing = (
            total_time_of_writing / self.__time_dict["number of chunks"]
        )

        # measure processing total time and avg time
        total_time_of_processing = (
            sum(self.__time_dict["reading"])
            + sum(self.__time_dict["filtering"])
            + sum(self.__time_dict["writing"])
        )
        avg_time_of_processing = (
            total_time_of_processing / self.__time_dict["number of chunks"]
        )

        data = {
            "D.frame size": self.__time_dict["chunksize"],
            "avg_Reading Time": avg_time_of_read,
            "avg_filtering Time": avg_time_of_filter,
            "Total_processing_Time": total_time_of_processing,
            "avg_processing_Time": avg_time_of_processing,
        }

        headers = list(data.keys())
        # check if the file exists
        try:
            workbook = openpyxl.load_workbook("./output/output.xlsx")
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.create_sheet(sheet_name)
                worksheet.append(headers)
        except FileNotFoundError:
            # create a new workbook and worksheet if the file doesn't exist
            workbook = openpyxl.Workbook()
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(headers)

        # delete the default 'Sheet' worksheet
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        # append data to the worksheet
        values = [data[header] for header in headers]
        worksheet.append(values)
        # save the workbook
        workbook.save("./output/output.xlsx")

    def start_filtering(self):
        elapsed_time = elapsed(self.__time_dict["start_time"])
        logging.info(
            "%s  consumer:start filtering chunks using %s.",
            elapsed_time,
            self.__text_filter,
        )
        self.__text_filter.prepare()
        while True:
            if self.__input_queue.empty():
                continue
            if (tuple := self.__input_queue.get()) is None:
                break
            # filtering
            start_time = time.time()
            healthy_chunk, unhealthy_chunk = self.__text_filter.filter(tuple[1])
            end_time = time.time()

            # add the time taken to filter the chunk
            self.__time_dict_lock.acquire()
            self.__time_dict["filtering"].append((tuple[0] + 1, end_time - start_time))
            self.__time_dict["number_of_healthy"] += healthy_chunk.shape[0]
            self.__time_dict["number_of_unhealthy"] += unhealthy_chunk.shape[0]
            self.__time_dict_lock.release()

            elapsed_time = elapsed(self.__time_dict["start_time"])
            logging.info(
                "%s  Consumer: finish filtering chunk number %s",
                elapsed_time,
                tuple[0] + 1,
            )

    def run(self):

        self.start_filtering()

        elapsed_time = elapsed(self.__time_dict["start_time"])
        logging.info(
            "%s  Consumer:starting write excel and csv files statistics ....",
            elapsed_time,
        )
        excel_thread = threading.Thread(
            target=self.excel_writer, args=(self.__text_filter.__repr__(),)
        )
        csv_thread = threading.Thread(
            target=self.write_csv_statistics, args=("individual_statistics.csv",)
        )

        excel_thread.start()
        csv_thread.start()
        excel_thread.join()
        csv_thread.join()

        elapsed_time = elapsed(self.__time_dict["start_time"])
        logging.info(
            "%s  Consumer:finish of write excel and csv files statistics .",
            elapsed_time,
        )
